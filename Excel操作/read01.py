# coding:utf8

"""
workbook 工作簿
worksheet 表单
row column cell 行 列 单元格
"""

import openpyxl

wb = openpyxl.load_workbook("example.xlsx") # 打开excel文件 工作簿类型的对象

# print(wb.get_sheet_by_name('Sheet1'))
# 从工作簿获取工作表单worksheet
print("第1个sheet名字",wb.sheetnames[0])
print("所有表单名字",wb.sheetnames) # 返回list

# 遍历工作簿，打印每个表单的名字
# for sheet in wb:
#     print(sheet.title)

# 创建一个表单名字为mySheet
# mySheet = wb.create_sheet('mySheet')
# print(wb.sheetnames) # 打印工作簿下所有表单名字
# print(mySheet) # 表单对象
# print(mySheet['A1']) # 定位A1单元格
# print(mySheet['A1'].value) # 取A1单元格的值

# 获取表单对象的两种方式：wb.get_sheet_by_name('sheet')与wb['sheet1']
# sheet3 = wb.get_sheet_by_name('Sheet3')
# print(type(sheet3)) # 表单对象
# print(sheet3)
#
# sheet4 = wb['mySheet']
# print(sheet4)
# print(sheet4[1])
#==============================#

# 操作表单的单元格，每个表单都有一个活跃的表单active
ws = wb.active  # Sheet1为当前活跃的表单
# print("活跃的表单对象",ws)
# print("定位A1单元格",ws['A1'])
# print("取A1单元格的值",ws['A1'].value)

# 打印行和列,引用单元格的方法：
c = ws['B1']
print("Row {} Column {} is {}".format(c.row, c.column, c.value)) # Row 1 Column B is 性别
print("Cell {} is {}".format(c.coordinate, c.value)) # Cell B1 is 性别
print(ws.cell(row=1, column=2))
print(ws.cell(row=1, column=2).value)

# print("---------------")
# 可以用for循环控制
# for i in range(1, 8):
#     for j in range(1, 4):
#         print(ws.cell(row=i, column=j).value)

# print("---------------")
# # 取C列所有数据
# colC = ws['C']
# print(colC)
# print(colC[2].value) # 索引是从0开始的，行标示从1开始的

# print("---------------")
# # 取第六行所有数据
# row6 = ws[6]
# print(row6)
# print(row6[2].value)

print("------列---------")
# 行列切片
"""列变化，先遍历B列所有单元格，"""
col_range = ws['B:C']
row_range = ws[2:6]

for col in col_range:
    for cell in col:
        print(cell.value)

print("---------行------")
for row in row_range:
    for cell in row:
        print(cell.value)
# ws.iter_rows方法
for row in ws.iter_rows(min_row=1, max_row=2, max_col=2):
    for cell in row:
        print(cell)

# 列和行 生成器
print(ws.rows)
print(tuple(ws.rows)[1])
print(tuple(ws.rows)[1][1])
print(tuple(ws.rows)[1][1].value)

print("---------单元格切片------")
# 单元格切片
cell_range = ws['A1:C3']

for rowofCell in cell_range:
    for cell in rowofCell:
        print(cell.coordinate, cell.value) # cell.coordinate打印行和列如A1

# 打印多少行多少列 7行三列
print('{}*{}'.format(ws.max_row, ws.max_column))

# 字母标转换成数字,数字转换成字母
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(2))
print(get_column_letter(47))
print(column_index_from_string('A'))
print(column_index_from_string('AB'))